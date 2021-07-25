# Install openpyxl, xlrd, pandas, XlsxWriter, seaborn, sty in your environment or using terminal
# importing the modules
import math
import os

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import seaborn as sn
# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
from plotly.subplots import make_subplots

# # The next section in this script needs to be assigned before running the code
# Assign spreadsheet filename to `fluor_file`
fluor_file = 'experiment_1.xlsx'

# dimensions of the plate
# given below for a 384 well plate (16x24)
total_rows = 16
total_columns = 24

# fluorescence reading labels as mentioned in the output file
labels = ["sfGFP20", "MGA80", "sfGFP40"]

# Active sheet in the output file
sheet_name = "Result sheet"


# Actual Code starts: user need not change any part after this

def fluorescence(flag, end_flag, name_sheet, file_output):
    # Load in the workbook
    wb = load_workbook(file_output)

    # define the active sheet by sheet_name
    sheet = wb[name_sheet]

    # getting the last row and column in the Sheet
    end_col = sheet.max_column

    # # storing and returning the fluorescence data for the mentioned fluorescence label
    time = []
    for i in range(2, end_col + 1):
        time.append(sheet.cell(row=2 + flag, column=i).value)
        # Saving the timepoints for fluorescence readout in an array (2nd row after the flag)

    # fluor_data is the dictionary where each item(well number) corresponds to an array(fluor_data of the well)
    fluor_data = {"Time": np.array(time)}
    data = []
    well_array = []  # array of well numbers
    for h in range(4 + flag, end_flag + 1):
        for i in range(2, end_col + 1):
            if isinstance(sheet.cell(row=h, column=i).value, int):
                # Stores the correct fluorescence readout ONLY IF an INTEGER value
                data.append(sheet.cell(row=h, column=i).value)
            else:
                # Stores numpy nan value if not an integer
                value = np.nan
                data.append(value)
        # saving the array of fluorescence data with well number as the key
        fluor_data[sheet.cell(row=h, column=1).value] = np.array(data)
        # Saving the well numbers in a separate array
        well_array.append(sheet.cell(row=h, column=1).value)
        data = []
    return fluor_data, well_array


# # For all the fluorescent labels this function plots the heatmap of Maxima, Minima, Endpoint, Slope of Trendline
# # in the plate grid format
def output_plots(tot_rows, tot_cols, fluor_labels, name_sheet, file_output):
    # creating a folder to store output plots if not already there
    if os.path.isdir("output_plots") is False:
        os.makedirs("output_plots")
    
    # Load in the workbook
    wb = load_workbook(file_output)

    # define the active sheet by sheet_name
    sheet = wb[name_sheet]

    # getting the last row and column in the Sheet
    end_row = sheet.max_row

    # getting the index for the row number in the output file where the block of fluorescence data corresponding to each label starts
    flags = {}
    r = 1
    while r < end_row + 1:
        if sheet.cell(row=r, column=1).value in fluor_labels:
            flags[sheet.cell(row=r, column=1).value] = r
        r += 1

    # getting the index for the row number in the output file where the block of fluorescence data corresponding to each label ends
    # this is based on the fact that the block of fluorescence data ends with a blank row (blank first cell)
    end_flags = {}
    lab = 0
    for f in fluor_labels:
        for i in range(flags[f], end_row + 2):
            if sheet.cell(row=i, column=1).value is None:
                if lab == 0:
                    end_flags[f] = i - 1
                    lab = 1
        lab = 0

    # major and minor axes for the row and column indexes of the plate respectively
    minor_axis, major_axis = plate_axes(tot_cols, tot_rows)

    # # The next whole chunk of code will be repeated for all the mentioned labels
    for f in fluor_labels:
        fluor_data, well_array = fluorescence(flags[f], end_flags[f], name_sheet, file_output)

        tag = ["Maxima", "Minima", "Endpoint", "Slope of Trendline"]  # features to be plotted

        # output data is the dictionary where each item(feature) corresponds to a dataframe(corresponding feature value in each well)
        output_data = {}
        data_max = []  # 2D array: list of rows(max values)
        data_min = []  # 2D array: list of rows(min values)
        data_end = []  # 2D array: list of rows(endpoint values)
        data_slope = []  # 2D array: list of rows(slope values)
        dataset_max = []  # 1D array containing max values in each row
        dataset_min = []  # 1D array containing min values in each row
        dataset_end = []  # 1D array containing endpoint values in each row
        dataset_slope = []  # 1D array containing slope values in each row

        time_filter = []
        fluor_data_filter = []

        for i in range(0, tot_rows):
            for j in range(0, tot_cols):
                string = major_axis[i] + str(minor_axis[j])
                if string in well_array:
                    maxima, minima, endpt = features(fluor_data[string])
                    dataset_max.append(maxima)
                    dataset_min.append(minima)
                    dataset_end.append(endpt)
                    if np.isnan(fluor_data[string]).any():
                        for x in range(0, len(fluor_data["Time"])):
                            if math.isnan(fluor_data[string][x]):
                                pass
                            else:  # filtering only non nan values and corresponding timepoints for slope calculation
                                time_filter.append(fluor_data["Time"][x])
                                fluor_data_filter.append(fluor_data[string][x])
                        z = np.polyfit(np.array(time_filter), np.array(fluor_data_filter), 1)
                    else:  # No filtering is required if neither of the values is nan
                        z = np.polyfit(fluor_data["Time"], fluor_data[string], 1)
                    dataset_slope.append(z[0])
                if string not in well_array:  # taking fluorescence value as 0 for wells which did not give a fluorescence readout
                    dataset_max.append(np.nan)
                    dataset_min.append(np.nan)
                    dataset_end.append(np.nan)
                    dataset_slope.append(np.nan)
            data_max.append(dataset_max)
            data_min.append(dataset_min)
            data_end.append(dataset_end)
            data_slope.append(dataset_slope)
            dataset_max = []
            dataset_min = []
            dataset_end = []
            dataset_slope = []
            time_filter = []
            fluor_data_filter = []
        output_data["Maxima"] = pd.DataFrame(data_max, index=major_axis, columns=minor_axis)
        output_data["Minima"] = pd.DataFrame(data_min, index=major_axis, columns=minor_axis)
        output_data["Endpoint"] = pd.DataFrame(data_end, index=major_axis, columns=minor_axis)
        output_data["Slope of Trendline"] = pd.DataFrame(data_slope, index=major_axis, columns=minor_axis)

        # generating a heatmap of the different features(tags) in the plate format using seaborn        
        for x in tag:
            grid_kws = {"height_ratios": (.9, .05), "hspace": .3}
            fig, (ax, cbar_ax) = plt.subplots(2, gridspec_kw=grid_kws)
            svm = sn.heatmap(output_data[x], linewidths=0.5, linecolor='lavender', cmap=fluor_colorscale(f), ax=ax, cbar_ax=cbar_ax,
                             cbar_kws={"orientation": "horizontal"})
            figure = svm.get_figure()
            figure.savefig('output_plots\\' + x + '_' + f + '.png', dpi=400)
    return

# classifying and defining the colorscale based on the fluorescent protein
def fluor_colorscale(text):
    if text.find("GFP") != -1:
        return 'BuGn'
    elif text.find("MGA") != -1:
        return 'OrRd'
    elif text.find("RFP") != -1:
        return 'PuRd'
    else:
        return 'YlGnBu'

# classifying and defining the solid linecolor based on the fluorescent protein
def fluor_linecolor(text):
    if text.find("GFP") != -1:
        return 'green'
    elif text.find("MGA") != -1:
        return 'red'
    elif text.find("RFP") != -1:
        return 'crimson'
    else:
        return 'cyan'

# computing the maxima, minima and endpoint value of an array
def features(arr):
    maxima = max(el for el in arr if el is not np.nan)
    minima = min(el for el in arr if el is not np.nan)
    endpt = arr[-1]
    return maxima, minima, endpt

# major and minor axes for the row and column indexes of the plate respectively
def plate_axes(tot_cols, tot_rows):
    minor_axis = []
    for i in range(1, tot_cols + 1):
        minor_axis.append(i)
    major_axis = []
    alpha = 'A'
    beta = ''
    for i in range(1, tot_rows + 1):
        if i == 27:
            alpha = 'A'
            beta = 'A'
        elif i > 27:
            if i % 26 == 1:
                alpha = 'A'
                beta = chr(ord(beta) + 1)
        major_axis.append(beta + alpha)
        alpha = chr(ord(alpha) + 1)
    return minor_axis, major_axis

# # Displaying the Kinetic traces in the multi-well plate format
def kinetic_trace(tot_rows, tot_cols, fluor_labels, name_sheet, file_output):
    # Load in the workbook
    wb = load_workbook(file_output)

    # define the active sheet by sheet_name
    sheet = wb[name_sheet]

    # getting the last row and column in the Sheet
    end_row = sheet.max_row

    # getting the index for the row number in the output file where the block of fluorescence data corresponding to each label starts
    flags = {}
    r = 1
    while r < end_row + 1:
        if sheet.cell(row=r, column=1).value in fluor_labels:
            flags[sheet.cell(row=r, column=1).value] = r
        r += 1

    # getting the index for the row number in the output file where the block of fluorescence data corresponding to each label ends
    # this is based on the fact that the block of fluorescence data ends with a blank row (blank first cell)
    end_flags = {}
    lab = 0
    for f in fluor_labels:
        for i in range(flags[f], end_row + 2):
            if sheet.cell(row=i, column=1).value is None:
                if lab == 0:
                    end_flags[f] = i - 1
                    lab = 1
        lab = 0

    # major and minor axes for the row and column indexes of the plate respectively
    minor_axis, major_axis = plate_axes(tot_cols, tot_rows)

    # pandas DataFrame of the well numbers
    well_data = []
    subdata = []
    for i in major_axis:
        for j in minor_axis:
            string = i + str(j)
            subdata.append(string)
        well_data.append(subdata)
        subdata = []
    well_number = pd.DataFrame(well_data, index=major_axis, columns=minor_axis)

    # # The next whole chunk of code will be repeated for all the mentioned labels
    for f in fluor_labels:
        fluor_data, well_array = fluorescence(flags[f], end_flags[f], name_sheet, file_output)

        # Plotting fluor_data vs time in a grid (given in multi-plate format)
        x = tot_cols
        y = tot_rows
        fig = make_subplots(rows=y, cols=x, column_titles=minor_axis, shared_yaxes=True,
                            shared_xaxes=True, horizontal_spacing=0.01, vertical_spacing=0.01)
        r = 1
        plot_max = []  # array of the (max fluor val + max SD) for all the wells
        # (this will help in fixing the y axis range of the fluorescence vs time plot)
        
        for j in major_axis:
            for i in minor_axis:
                if well_number[i][j] in well_array:
                    fig.add_trace(go.Scatter(x=fluor_data["Time"], y=fluor_data[well_number[i][j]], name=well_number[i][j],
                                             hovertext=well_number[i][j], line=dict(color=fluor_linecolor(f)), hoverinfo='x+y+text', showlegend=False),
                                  row=r, col=i)
                    plot_max.append(max(el for el in fluor_data[well_number[i][j]] if el is not np.nan))
                else:
                    fig.add_trace(go.Scatter(x=fluor_data["Time"], y=[0], name=well_number[i][j], hovertext=well_number[i][j],
                                             hoverinfo='text', showlegend=False), row=r, col=i)
            r += 1

        fig.update_yaxes(range=[0, max(plot_max)])  # fixes the range of yaxis
        fig.update_xaxes(range=[0, max(fluor_data["Time"])])  # fixes the range of xaxis
        for i in range(0, y):
            fig.update_yaxes(title_text=major_axis[i], row=i+1, col=1)
        fig.update_layout(height=1500, width=2500, title_text="Fluorescence reading vs Time: " + f)
        fig.show()

    return

# Calling the function: 'output_plots' and 'kinetic_trace'
output_plots(total_rows, total_columns, labels, sheet_name, fluor_file)
kinetic_trace(total_rows, total_columns, labels, sheet_name, fluor_file)

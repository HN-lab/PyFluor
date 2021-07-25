# Install openpyxl, xlrd, pandas, XlsxWriter, seaborn, sty in your environment or using terminal
# importing the modules
import numpy as np
import pandas as pd
import plotly.graph_objects as go
# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
from plotly.subplots import make_subplots

# # The next section in this script needs to be assigned before running the code
# Assign spreadsheet filename to `file`
file_in = 'design_table_1.xlsx'
file_out = 'experiment_1.xlsx'

# dimensions of the plate
# given below for a 384 well plate (16x24)
total_rows = 16
total_columns = 24

# fluorescence reading labels as mentioned in the output file
labels = ["sfGFP20", "MGA80", "sfGFP40"]

# Active sheet in the output file
sheet_name = "Result sheet"

# features to plot against each other
expt = ["DNA", "Maltose"]
# Are the other reagents are also varying?(comment the line which does not apply)
answer = "no"
# answer = "yes"


# Actual Code starts: user need not change any part after this

def volume_dict(tot_rows, tot_cols, file_input):
    # Load in the workbook
    wb = load_workbook(file_input)

    # define the different sheets by name
    sheet = wb['Sheet1']

    # last row in Sheet1
    end = sheet.max_row

    # major and minor axes for the row and column indexes of the plate respectively
    minor_axis, major_axis = plate_axes(tot_cols, tot_rows)

    i = 1
    items = []  # list of reagents
    while i < end:
        items.append(sheet.cell(row=i, column=tot_cols + 2).value)
        i += (tot_rows + 1)

    # volume data is the dictionary where each item(reagent) corresponds to a dataframe(volume of reagent to be added to the plate)
    volume_data = {}
    data = []  # 2D array: list of rows
    dataset = []  # 1D array containing values in each row
    h = 0
    i = 1
    while i < end:
        k = 0
        while k < tot_rows:
            j = 2
            while j < tot_cols + 2:
                if sheet.cell(row=i + k + 1, column=j).value is None:
                    dataset.append(np.nan)
                else:
                    dataset.append(sheet.cell(row=i + k + 1, column=j).value)
                j += 1
            data.append(dataset)
            dataset = []
            k += 1
        volume_data[items[h]] = pd.DataFrame(data, index=major_axis, columns=minor_axis)
        data = []
        i += (tot_rows + 1)
        h += 1
    # returning the volume data dictionary
    return volume_data, items


def fluorescence(flag, end_flag, name_sheet, file_output):
    # Load in the workbook
    wb = load_workbook(file_output)

    # define the active sheet by sheet_name
    sheet = wb[name_sheet]

    # getting the last row and column in the Sheet
    end_col = sheet.max_column

    # # storing and returning the fluorescence data for the mentioned fluorescence label
    time = []
    for i in range(2, end_col + 1):
        time.append(sheet.cell(row=2 + flag, column=i).value)
        # Saving the timepoints for fluorescence readout in an array (2nd row after the flag)

    # fluor_data is the dictionary where each item(well number) corresponds to an array(fluor_data of the well)
    fluor_data = {"Time": np.array(time)}
    data = []
    well_array = []  # array of well numbers
    for h in range(4 + flag, end_flag + 1):
        for i in range(2, end_col + 1):
            if isinstance(sheet.cell(row=h, column=i).value, int):
                # Stores the correct fluorescence readout ONLY IF an INTEGER value
                data.append(sheet.cell(row=h, column=i).value)
            else:
                # Stores numpy nan value if not an integer
                value = np.nan
                data.append(value)
        # saving the array of fluorescence data with well number as the key
        fluor_data[sheet.cell(row=h, column=1).value] = np.array(data)
        # Saving the well numbers in a separate array
        well_array.append(sheet.cell(row=h, column=1).value)
        data = []
    return fluor_data, well_array


# # For all the fluorescent labels this function plots experiment specific plots: one reagent vs another (eg. Maltose vs DNA)
# # The two reagents that are to be plotted against each other are to be mentioned in the beginning of the code in the "expt" array

def expt_plots(tot_rows, tot_cols, fluor_labels, expt_tags, name_sheet, file_input, file_output, answer_tags):
    # Load in the workbook
    wb = load_workbook(file_output)

    # define the active sheet by sheet_name
    sheet = wb[name_sheet]

    # getting the last row and column in the Sheet
    end_row = sheet.max_row

    # getting the index for the row number in the output file where the block of fluorescence data corresponding to each label starts
    flags = {}
    r = 1
    while r < end_row + 1:
        if sheet.cell(row=r, column=1).value in fluor_labels:
            flags[sheet.cell(row=r, column=1).value] = r
        r += 1

    # getting the index for the row number in the output file where the block of fluorescence data corresponding to each label ends
    # this is based on the fact that the block of fluorescence data ends with a blank row (blank first cell)
    end_flags = {}
    lab = 0
    for f in fluor_labels:
        for i in range(flags[f], end_row + 2):
            if sheet.cell(row=i, column=1).value is None:
                if lab == 0:
                    end_flags[f] = i - 1
                    lab = 1
        lab = 0

    # major and minor axes for the row and column indexes of the plate respectively
    minor_axis, major_axis = plate_axes(tot_cols, tot_rows)

    # pandas DataFrame of the well numbers
    well_data = []
    subdata = []
    for i in major_axis:
        for j in minor_axis:
            string = i + str(j)
            subdata.append(string)
        well_data.append(subdata)
        subdata = []
    well_number = pd.DataFrame(well_data, index=major_axis, columns=minor_axis)

    # making a list of the distinct volumes of the features to be plotted against each other (given by the expt array)
    volume_data, items = volume_dict(tot_rows, tot_cols, file_input)
    count = {}
    element = {}
    for x in volume_data.keys():
        count[x] = 0
        element[x] = []
        for i in range(1, tot_cols + 1):
            for el in volume_data[x][i]:
                if el not in element[x]:
                    if np.isnan(el):
                        pass
                    else:
                        element[x].append(el)
                        count[x] += 1
        print("No.of.unique values for", x, ": ", count[x])
        print("unique values :", element[x])

    # giving labels to the list of the distinct volumes of the features to be plotted against each other (given by the element array)
    index = {}
    index_array = []
    for el in volume_data.keys():
        for i in range(0, count[el]):
            index_array.append(el + " volume= " + str(element[el][i]))
        index[el] = index_array
        index_array = []

    # array of reagents other than those mentioned in expt array
    other_reagents = [el for el in set(volume_data.keys()) - set(expt_tags)]

    # # The next whole chunk of code will be repeated for all the mentioned labels
    for f in fluor_labels:
        fluor_data, well_array = fluorescence(flags[f], end_flags[f], name_sheet, file_output)

        # Defining an expt specific grid (given by the two reagents to be plotted against each other)
        grid = []
        subgrid = []
        well = []
        for j in range(0, len(element[expt_tags[1]])):
            for k in range(0, len(element[expt_tags[0]])):
                for h in major_axis:
                    for i in minor_axis:
                        if well_number[i][h] in well_array:
                            if element[expt_tags[0]][k] == volume_data[expt_tags[0]][i][h] and element[expt_tags[1]][j] == volume_data[expt_tags[1]][i][h]:
                                well.append(well_number[i][h])
                subgrid.append(well)
                well = []
            grid.append(subgrid)
            subgrid = []

        expt_grid = pd.DataFrame(grid, index=index[expt_tags[1]], columns=index[expt_tags[0]])

        # Dictionary carrying information about all the reagents in the well except the ones mentioned in the Expt Array
        other_r_well = {}
        other_r_well_data = []
        for h in major_axis:
            for i in minor_axis:
                if well_number[i][h] in well_array:
                    for item in other_reagents:
                        if np.isnan(volume_data[item][i][h]):
                            pass
                        else:
                            other_r_well_data.append(item + " volume= " + str(volume_data[item][i][h]))
                    other_r_well[well_number[i][h]] = other_r_well_data
                    other_r_well_data = []

        new_grid = expt_grid.copy()
        # # The next chunk of code will be iterated over all the columns in expt_grid if the other reagents are also varying
        if answer_tags == "yes":
            for col_old in expt_grid.columns:
                # To recognize replicates
                new_col_data = []
                wells = [item for innerlist in expt_grid[col_old] for item in innerlist]
                other_r_data = [other_r_well[x] for x in wells]
                other_r_data = [list(item) for item in set(tuple(row) for row in other_r_data)]
                new_col_data.extend([col_old for _ in range(len(other_r_data))])

                col_r = {}
                col_r_data = []
                for y in other_r_data:
                    col_r[str(y)] = []
                for y in other_r_data:
                    for x in expt_grid[col_old]:
                        for el in x:
                            if other_r_well[el] == y:
                                col_r_data.append(el)
                        col_r[str(y)].append(col_r_data)
                        col_r_data = []

                # inserting columns in new_grid
                for i in range(0, len(other_r_data)):
                    new_grid.insert(new_grid.columns.get_loc(col_old) + 1,
                                    str(new_col_data[i]) + " " + str(other_r_data[i]), col_r[str(other_r_data[i])],
                                    True)
                del new_grid[col_old]
                
        # # Plotting fluor_data vs time in a grid (given by the counts of distinct volumes of the expt reagents to be plotted against each other)
        x = len(new_grid.columns)
        y = len(new_grid.index)
        fig = make_subplots(rows=y, cols=x, shared_yaxes=True, shared_xaxes=True, horizontal_spacing=0.01, vertical_spacing=0.03)
        r = len(new_grid.index)
        c = 1
        plot_max = []  # array of the (max fluor val + max SD) for all the wells
        # (this will help in fixing the y axis range of the fluorescence vs time plot)
        well_string = []
        text_string = []
        # the above two arrays help in creating hovertext(well numbers) for the heatmap
        expt_maxima_grid = new_grid.copy()  # copying the skeleton to making an expt_grid of maxima values
        for j in new_grid.index:
            for i in new_grid:
                if len(new_grid[i][j]) > 0:
                    multiple_lists = []
                    n = 0
                    for el in new_grid[i][j]:
                        multiple_lists.append(fluor_data[el])
                        n += 1
                    datapoints = np.mean(multiple_lists, axis=0)
                    std_dev = np.std(multiple_lists, axis=0)
                    well_labels = ""
                    for el in new_grid[i][j]:
                        well_labels = el + "  " + well_labels
                    text_string.append(well_labels)
                    fig.add_trace(go.Scatter(name='Upper Bound', x=fluor_data["Time"], y=np.add(datapoints, std_dev),
                                             marker=dict(color="#444"),
                                             line=dict(width=0), showlegend=False, hoverinfo='skip'), row=r, col=c)
                    fig.add_trace(go.Scatter(name='Lower Bound', x=fluor_data["Time"], y=np.subtract(datapoints, std_dev),
                                             marker=dict(color="#444"),
                                             line=dict(width=0), fillcolor='rgba(68, 68, 68, 0.3)', fill='tonexty',
                                             showlegend=False, hoverinfo='skip'), row=r, col=c)
                    fig.add_trace(go.Scatter(x=fluor_data["Time"], y=datapoints, name=well_labels,
                                             line=dict(color=fluor_linecolor(f)),
                                             hoverinfo='all', hovertext=i + ", " + j, showlegend=False), row=r, col=c)

                    expt_maxima_grid[i][j] = max(el for el in datapoints if el is not np.nan)
                    sd_max = max(el for el in std_dev if el is not np.nan)
                    plot_max.append(expt_maxima_grid[i][j] + sd_max)
                else:
                    text_string.append("No Wells")
                    fig.add_trace(go.Scatter(x=fluor_data["Time"], y=[0 for _ in fluor_data["Time"]],
                                             line=dict(color=fluor_linecolor(f)), name="No Wells", hoverinfo='text+name',
                                             hovertext=i + ", " + j, showlegend=False), row=r, col=c)
                c += 1
            r -= 1
            c = 1
            well_string.append(text_string)
            text_string = []

        fig.update_yaxes(range=[0, max(plot_max)])  # fixes the range of yaxis
        for j in range(0, x):
            fig.update_xaxes(title_text=new_grid.columns[j], row=y, col=j + 1)
        for i in range(0, y):
            fig.update_yaxes(title_text=new_grid.index[i], row=y - i, col=1)
        fig.update_layout(height=1000, title_text="Fluorescence reading vs Time: " + f)
        fig.show()

        # Plotting a heatmap for expt_maxima_grid        
        expt_maxima_data = {'z': expt_maxima_grid.values.tolist(), 'x': expt_maxima_grid.columns.tolist(),
                            'y': expt_maxima_grid.index.tolist()}

        fig = go.Figure(data=go.Heatmap(expt_maxima_data, colorscale=fluor_colorscale(f), hovertext=well_string))
        fig.update_layout(title={
            'text': "Maximum Fluorescence reading with different volumes of " + expt_tags[0] + " and " + expt_tags[
                1] + ": " + f},
            autosize=True)
        fig.show()
    return


# classifying and defining the colorscale based on the fluorescent protein
def fluor_colorscale(text):
    if text.find("GFP") != -1:
        return 'BuGn'
    elif text.find("MGA") != -1:
        return 'OrRd'
    elif text.find("RFP") != -1:
        return 'PuRd'
    else:
        return 'YlGnBu'


# classifying and defining the solid linecolor based on the fluorescent protein
def fluor_linecolor(text):
    if text.find("GFP") != -1:
        return 'green'
    elif text.find("MGA") != -1:
        return 'red'
    elif text.find("RFP") != -1:
        return 'crimson'
    else:
        return 'cyan'


# computing the maxima, minima and endpoint value of an array
def features(arr):
    maxima = max(el for el in arr if el is not np.nan)
    minima = min(el for el in arr if el is not np.nan)
    endpt = arr[-1]
    return maxima, minima, endpt


# major and minor axes for the row and column indexes of the plate respectively

def plate_axes(tot_cols, tot_rows):
    minor_axis = []
    for i in range(1, tot_cols + 1):
        minor_axis.append(i)
    major_axis = []
    alpha = 'A'
    beta = ''
    for i in range(1, tot_rows + 1):
        if i == 27:
            alpha = 'A'
            beta = 'A'
        elif i > 27:
            if i % 26 == 1:
                alpha = 'A'
                beta = chr(ord(beta) + 1)
        major_axis.append(beta + alpha)
        alpha = chr(ord(alpha) + 1)
    return minor_axis, major_axis


# Calling the function: 'expt_plots'
expt_plots(total_rows, total_columns, labels, expt, sheet_name, file_in, file_out, answer)
